# Databricks notebook source
pip install office365-rest-python-client --trusted-host pypi.org --trusted-host files.pythonhosted.org openpyxl

# COMMAND ----------

# Databricks notebook / script
# Reads distinct (system_id, account_name) from prod_l2.services.tbl_ocs_ei_customer_account,
# wipes Sheet2 of the SharePoint workbook, writes the data fresh, and turns it into an
# Excel Table named "Table1". Then re-uploads the file back to SharePoint.

import io
import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from office365.runtime.auth.client_credential import ClientCredential
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File

# ---------------------------------------------------------------------------
# 1. Config — consider moving secrets to Databricks Secret Scopes instead of
#    hardcoding them, e.g. dbutils.secrets.get(scope="sp", key="client_secret")
# ---------------------------------------------------------------------------
tenant_id = ""
client_id = ""
client_secret = ""

site_url = "https://share.philips.com/sites/RI-CPMExcellence"
file_relative_url = "/sites/RI-CPMExcellence/Shared Documents/General/Dashboard/input/final_project_account_mapping_updated_logic.xlsx"

SHEET_NAME = "Sheet2"
TABLE_NAME = "Table1"

# ---------------------------------------------------------------------------
# 2. Read distinct values from the Spark table
# ---------------------------------------------------------------------------
query = """
    SELECT DISTINCT system_id, account_name
    FROM prod_l2.services.tbl_ocs_ei_customer_account
"""
pdf = spark.sql(query).toPandas()
pdf = pdf[["system_id", "account_name"]]

print(f"Fetched {len(pdf)} distinct rows from Spark table.")

# ---------------------------------------------------------------------------
# 3. Connect to SharePoint and download the workbook
# ---------------------------------------------------------------------------
credentials = ClientCredential(client_id, client_secret)
ctx = ClientContext(site_url).with_credentials(credentials)

response = File.open_binary(ctx, file_relative_url)
file_content = response.content

wb = openpyxl.load_workbook(io.BytesIO(file_content))

# ---------------------------------------------------------------------------
# 4. Get (or create) Sheet2, and wipe it completely
# ---------------------------------------------------------------------------
if SHEET_NAME in wb.sheetnames:
    ws = wb[SHEET_NAME]
    # Remove any existing tables on this sheet (avoids duplicate-name errors
    # and leftover table definitions pointing at stale ranges)
    for existing_table_name in list(ws.tables.keys()):
        del ws.tables[existing_table_name]
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
else:
    ws = wb.create_sheet(SHEET_NAME)

# ---------------------------------------------------------------------------
# 5. Write header + data starting at A1
# ---------------------------------------------------------------------------
headers = list(pdf.columns)
ws.append(headers)
for record in pdf.itertuples(index=False):
    ws.append(list(record))

n_rows = len(pdf)
n_cols = len(headers)
end_col_letter = get_column_letter(n_cols)
table_ref = f"A1:{end_col_letter}{n_rows + 1}"  # +1 for header row

# ---------------------------------------------------------------------------
# 6. Create Table1 over the freshly written range
# ---------------------------------------------------------------------------
excel_table = Table(displayName=TABLE_NAME, ref=table_ref)
excel_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(excel_table)

print(f"Created table '{TABLE_NAME}' over range {table_ref} on '{SHEET_NAME}'.")

# ---------------------------------------------------------------------------
# 7. Save workbook to an in-memory buffer
# ---------------------------------------------------------------------------
output_buffer = io.BytesIO()
wb.save(output_buffer)
output_buffer.seek(0)
file_bytes = output_buffer.read()

# ---------------------------------------------------------------------------
# 8. Upload the updated workbook back to the same SharePoint location
# ---------------------------------------------------------------------------
folder_url, file_name = file_relative_url.rsplit("/", 1)
target_folder = ctx.web.get_folder_by_server_relative_url(folder_url)
target_folder.upload_file(file_name, file_bytes).execute_query()

print(f"Successfully updated '{file_name}' on SharePoint ({SHEET_NAME} / {TABLE_NAME}).")